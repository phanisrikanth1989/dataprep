"""
FileOutputExcelComponent - Writes data to an Excel file with advanced formatting and structure options.

Talend equivalent: tFileOutputExcel
"""
from typing import Dict, Any, List, Optional, Union
import pandas as pd
import openpyxl
import os
import logging

from ...base_component import BaseComponent
from ...component_registry import REGISTRY
from ...exceptions import FileOperationError, ComponentExecutionError, ConfigurationError

logger = logging.getLogger(__name__)


@REGISTRY.register("FileOutputExcel", "tFileOutputExcel")
class FileOutputExcel(BaseComponent):
    """
    Writes DataFrame data to an Excel file with support for multiple sheets and formatting options.

    Configuration:
        filename (str): Path to the Excel file to create/write to. Required.
        sheetname (str): Name of the Excel sheet to write to. Default: 'Sheet1'
        includeheader (bool): Whether to include column headers in output. Default: False
        append_file (bool): Whether to append to existing file or create new. Default: False
        create (bool): Whether to create output directory if it doesn't exist. Default: True
        encoding (str): File encoding (for compatibility, not used in Excel). Default: 'UTF-8'
        die_on_error (bool): Whether to fail job on error. Default: True

    Inputs:
        main: Primary input DataFrame containing data to write

    Outputs:
        None: This component does not produce output data

    Statistics:
        NB_LINE: Total rows processed
        NB_LINE_OK: Rows successfully written
        NB_LINE_REJECT: Rows rejected/skipped

    Example configuration:
        {
            "filename": "/path/to/output.xlsx",
            "sheetname": "DataSheet",
            "includeheader": true,
            "append_file": false,
            "create": true,
            "die_on_error": true
        }

    Notes:
        - Supports both DataFrame and list-of-dict input formats
        - Automatically filters out empty rows (all null/empty values)
        - Preserves column order from output schema if defined
        - Creates output directory if it doesn't exist
    """

    # Class constants
    DEFAULT_SHEET_NAME = 'Sheet1'
    DEFAULT_ENCODING = 'UTF-8'

    def _validate_config(self) -> None:
        """Validate component configuration.

        Raises:
            ConfigurationError: If required config is missing or has wrong shape.
        """
        if not self.config.get('filename'):
            raise ConfigurationError(f"[{self.id}] Missing required config 'filename'")
        if not isinstance(self.config['filename'], str):
            raise ConfigurationError(f"[{self.id}] Config 'filename' must be a non-empty string")
        if 'sheetname' in self.config and not isinstance(self.config['sheetname'], str):
            raise ConfigurationError(f"[{self.id}] Config 'sheetname' must be a string")
        for field_name in ('includeheader', 'append_file', 'append_sheet', 'create'):
            if field_name in self.config and not isinstance(self.config[field_name], bool):
                raise ConfigurationError(
                    f"[{self.id}] Config '{field_name}' must be a boolean"
                )

    def _process(self, input_data: Union[Dict[str, Any], pd.DataFrame, None] = None) -> Dict[str, Any]:
        """
        Write data to Excel file based on configuration.

        Args:
            input_data: Input data - can be DataFrame, dict containing DataFrames, or None

        Returns:
            Dictionary with execution statistics

        Raises:
            FileOperationError: If file operations fail
            ComponentExecutionError: If processing fails
        """
        # Handle empty input
        if input_data is None:
            logger.warning(f"[{self.id}] Empty input received")
            self._update_stats(0, 0, 0)
            return {'main': None, 'stats': self.stats}

        rows_in = 0
        rows_out = 0
        rows_rejected = 0

        logger.info(f"[{self.id}] Processing started")

        try:
            # Extract configuration with defaults
            filename = self.config['filename']
            sheet_name = self.config.get('sheetname', self.DEFAULT_SHEET_NAME)

            # Clean sheet name - remove quotes if present
            if sheet_name.startswith("'") and sheet_name.endswith("'"):
                sheet_name = sheet_name[1:-1]
            elif sheet_name.startswith('"') and sheet_name.endswith('"'):
                sheet_name = sheet_name[1:-1]

            include_header = self.config.get('includeheader', False)
            append_file = self.config.get('append_file', False)
            create_file = self.config.get('create', True)
            die_on_error = self.config.get('die_on_error', True)

            logger.info(f"[{self.id}] Writing to file: {filename}, sheet: {sheet_name}")

            # Ensure the output directory exists
            output_dir = os.path.dirname(filename)
            if create_file and output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                    logger.info(f"[{self.id}] Created output directory: {output_dir}")
                except OSError as e:
                    error_msg = f"Failed to create output directory: {output_dir}"
                    logger.error(f"[{self.id}] {error_msg}: {e}")
                    if die_on_error:
                        raise FileOperationError(f"[{self.id}] {error_msg}") from e
                    else:
                        self._update_stats(0, 0, 0)
                        return {'main': None, 'stats': self.stats}

            # Load or create workbook
            try:
                if append_file and os.path.exists(filename):
                    workbook = openpyxl.load_workbook(filename)
                    logger.info(f"[{self.id}] Loaded existing workbook: {filename}")
                else:
                    workbook = openpyxl.Workbook()
                    # Remove default sheet if it exists and we're creating a custom named sheet
                    if 'Sheet' in workbook.sheetnames and sheet_name != 'Sheet':
                        default_sheet = workbook['Sheet']
                        workbook.remove(default_sheet)
                        logger.info(f"[{self.id}] Removed default 'Sheet' to avoid multiple sheets")
            except Exception as e:
                error_msg = f"Failed to load/create workbook: {filename}"
                logger.error(f"[{self.id}] {error_msg}: {e}")
                if die_on_error:
                    raise FileOperationError(f"[{self.id}] {error_msg}") from e
                else:
                    self._update_stats(0, 0, 0)
                    return {'main': None, 'stats': self.stats}

            # Get or create sheet
            try:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    logger.info(f"[{self.id}] Using existing sheet: {sheet_name}")
                else:
                    sheet = workbook.create_sheet(sheet_name)
                    logger.info(f"[{self.id}] Created new sheet: {sheet_name}")
            except Exception as e:
                error_msg = f"Failed to create/access sheet: {sheet_name}"
                logger.error(f"[{self.id}] {error_msg}: {e}")
                if die_on_error:
                    raise ComponentExecutionError(self.id, error_msg, e) from e
                else:
                    self._update_stats(0, 0, 0)
                    return {'main': None, 'stats': self.stats}

            # Handle DataFrame input properly - Auto-detect from tMap outputs
            main_data = None

            # Check if input_data is a DataFrame directly
            if isinstance(input_data, pd.DataFrame):
                main_data = input_data
                logger.info(f"[{self.id}] Using input DataFrame directly with {len(input_data)} rows")
            elif isinstance(input_data, dict):
                # Look for 'main' key first
                main_data = input_data.get('main')

                # If 'main' not found, try to get the first non-stats DataFrame
                if main_data is None:
                    logger.info(f"[{self.id}] 'main' key not found, searching for first DataFrame in input_data")
                    logger.debug(f"[{self.id}] Available keys: {list(input_data.keys())}")

                    # Look for first DataFrame (skip 'stats' and other metadata)
                    for key, value in input_data.items():
                        if key != 'stats' and hasattr(value, 'iterrows'):  # It's a DataFrame
                            main_data = value
                            logger.info(f"[{self.id}] Using DataFrame from key '{key}' with {len(value)} rows")
                            break

            # Convert DataFrame to list of dictionaries if needed
            rows = []
            column_names = []

            if isinstance(main_data, pd.DataFrame):  # It's a pandas DataFrame
                rows_in = len(main_data)
                logger.info(f"[{self.id}] Converting DataFrame with {rows_in} rows to records")

                # Apply date_pattern formatting on a working copy
                df_out = main_data.copy()
                df_out = self._apply_date_patterns(df_out)

                # PRIORITY 1: Use input schema column order if defined (output_schema is empty for sink components)
                _input_schema = getattr(self, "input_schema", None) or []
                if _input_schema:
                    column_names = [col_def['name'] for col_def in _input_schema]
                    logger.debug(f"[{self.id}] Using input schema column order: {column_names}")
                elif getattr(self, "output_schema", None):
                    column_names = [col_def['name'] for col_def in self.output_schema]  # type: ignore[attr-defined]
                    logger.debug(f"[{self.id}] Using output schema column order: {column_names}")
                else:
                    # PRIORITY 3: Fall back to DataFrame column order
                    column_names = list(df_out.columns)
                    logger.debug(f"[{self.id}] Using DataFrame column order: {column_names}")

                # Convert to records while preserving column order
                for _, pandas_row in df_out.iterrows():
                    # Create ordered dictionary using the defined column order
                    row_dict = {}
                    for col in column_names:
                        # Handle case where schema column might not exist in DataFrame
                        if col in df_out.columns:
                            row_dict[col] = pandas_row[col]
                        else:
                            row_dict[col] = None  # Default None for missing schema columns
                            logger.warning(f"[{self.id}] Column '{col}' from schema not found in DataFrame, defaulting to None")
                    rows.append(row_dict)

            elif isinstance(main_data, list):
                rows = main_data
                rows_in = len(rows)
                # Use schema if available, otherwise use first row keys
                _input_schema = getattr(self, "input_schema", None) or []
                if _input_schema:
                    column_names = [col_def['name'] for col_def in _input_schema]
                    logger.debug(f"[{self.id}] Using input schema column order for list data: {column_names}")
                elif getattr(self, "output_schema", None):
                    column_names = [col_def['name'] for col_def in self.output_schema]  # type: ignore[attr-defined]
                    logger.debug(f"[{self.id}] Using output schema column order for list data: {column_names}")
                else:
                    column_names = list(rows[0].keys()) if rows else []
                    logger.debug(f"[{self.id}] Using first row keys for column order: {column_names}")
            else:
                logger.warning(f"[{self.id}] No data to write or unsupported data format")
                rows = []
                column_names = []
                rows_in = 0

            # Filter out empty rows (rows where all values are null/empty)
            def is_non_empty_row(row):
                """Return True if row has at least one non-null, non-empty value."""
                for value in row.values():
                    try:
                        if pd.isna(value):
                            continue
                    except (TypeError, ValueError):
                        pass  # pd.isna can fail on non-scalar types; treat as non-empty
                    if isinstance(value, str) and not value.strip():
                        continue
                    return True
                return False

            non_empty_rows = [row for row in rows if is_non_empty_row(row)]
            rows_rejected = len(rows) - len(non_empty_rows)

            logger.info("[%s] Processing %d non-empty rows out of %d total rows",
                        self.id, len(non_empty_rows), rows_in)

            # Find the actual last row that contains data (openpyxl can report max_row=1 for
            # a freshly-created empty sheet — a "ghost" row — which would otherwise shift all
            # content down by one row and cause duplicate headers on subsequent appends).
            def _last_data_row(ws):
                """Return the index of the last row with at least one non-None cell, or 0."""
                top = ws.max_row or 0
                for row_idx in range(top, 0, -1):
                    if any(cell.value is not None for cell in ws[row_idx]):
                        return row_idx
                return 0

            last_data_row = _last_data_row(sheet)

            # Write header only when the sheet is truly empty (no prior data).
            should_write_header = include_header and bool(column_names) and (last_data_row == 0)
            if should_write_header:
                logger.debug("[%s] Writing header row (sheet is empty)", self.id)
            else:
                logger.debug("[%s] Skipping header write (last_data_row=%d, include_header=%s)",
                             self.id, last_data_row, include_header)

            # Determine write start position from FIRST_CELL_X/Y config (0-based → 1-based)
            try:
                start_col = int(self.config.get('first_cell_x') or '0') + 1
            except (ValueError, TypeError):
                start_col = 1
            try:
                start_row = int(self.config.get('first_cell_y') or '0') + 1
            except (ValueError, TypeError):
                start_row = 1

            # For append_file or append_sheet mode, start writing after last existing data row
            append_sheet = self.config.get('append_sheet', False)
            if (append_file or append_sheet) and last_data_row > 0:
                start_row = max(start_row, last_data_row + 1)

            current_row = start_row

            # Build per-column number formats for Decimal/float precision (applied to cells)
            col_formats = self._build_col_formats()

            def _clean_val(value):
                """Convert NaN/NaT to None so openpyxl writes blank cells (not 'nan' strings).

                Also converts ``decimal.Decimal`` (Arrow BigDecimal with scale-18) to
                native Python ``float`` so Excel stores a proper number rather than a
                string with 18 trailing zeros.
                """
                if value is None:
                    return None
                import decimal as _decimal
                if isinstance(value, _decimal.Decimal):
                    return float(value)
                try:
                    if pd.isna(value):
                        return None
                except (TypeError, ValueError):
                    pass
                return value

            if should_write_header:
                for col_idx, col_name in enumerate(column_names):
                    sheet.cell(row=current_row, column=start_col + col_idx).value = col_name  # type: ignore[union-attr]
                current_row += 1
                logger.debug(f"[{self.id}] Added header row at row {current_row - 1}, col {start_col}")

            # Write data rows using sheet.cell() for FIRST_CELL_X/Y positioning support
            rows_written = 0
            for row in non_empty_rows:
                if column_names:
                    row_values = [_clean_val(row.get(col)) for col in column_names]
                else:
                    row_values = [_clean_val(v) for v in row.values()]
                for col_idx, value in enumerate(row_values):
                    cell = sheet.cell(row=current_row, column=start_col + col_idx)
                    cell.value = value  # type: ignore[union-attr]
                    # Apply number format for Decimal/float precision columns
                    col_name = column_names[col_idx] if col_idx < len(column_names) else None
                    if col_name and col_name in col_formats:
                        cell.number_format = col_formats[col_name]
                current_row += 1
                rows_written += 1

            rows_out = rows_written

            # Auto-size columns if IS_ALL_AUTO_SZIE or per-column AUTO_SZIE_SETTING
            if self.config.get('is_all_auto_szie', False) and column_names:
                from openpyxl.utils import get_column_letter
                for col_idx_0, col_name in enumerate(column_names):
                    col_letter = get_column_letter(start_col + col_idx_0)
                    max_len = len(str(col_name)) if include_header else 0
                    for row in non_empty_rows:
                        max_len = max(max_len, len(str(row.get(col_name) or '')))
                    sheet.column_dimensions[col_letter].width = max_len + 2
                logger.debug(f"[{self.id}] Auto-sized {len(column_names)} columns")
            elif self.config.get('auto_szie_setting') and column_names:
                from openpyxl.utils import get_column_letter
                for col_setting in self.config['auto_szie_setting']:
                    col_name = str(col_setting)
                    if col_name in column_names:
                        col_letter = get_column_letter(start_col + column_names.index(col_name))
                        max_len = max(
                            (len(str(row.get(col_name) or '')) for row in non_empty_rows),
                            default=0,
                        )
                        sheet.column_dimensions[col_letter].width = max_len + 2

            # Formula recalculation (RECALCULATE_FORMULA)
            if self.config.get('recalculate_formula', False):
                workbook.calculation.calcMode = 'auto'
                logger.debug(f"[{self.id}] Formula recalculation mode set to auto")

            # Save workbook
            try:
                workbook.save(filename)
                workbook.close()
                logger.info(f"[{self.id}] Excel file written successfully: {filename}")
            except Exception as e:
                error_msg = f"Failed to save Excel file: {filename}"
                logger.error(f"[{self.id}] {error_msg}: {e}")
                if die_on_error:
                    raise FileOperationError(f"[{self.id}] {error_msg}") from e
                else:
                    self._update_stats(rows_in, 0, rows_rejected)
                    return {'main': None, 'stats': self.stats}

            # Delete empty file if no data rows written (DELETE_EMPTYFILE)
            if self.config.get('delete_empty_file', False) and rows_out == 0:
                try:
                    os.remove(filename)
                    logger.info(f"[{self.id}] Deleted empty output file: {filename}")
                except OSError as e:
                    logger.warning(f"[{self.id}] Could not delete empty file {filename}: {e}")

            # Update statistics
            self._update_stats(rows_in, rows_out, rows_rejected)
            logger.info(f"[{self.id}] Processing complete: in={rows_in}, out={rows_out}, rejected={rows_rejected}")

            return {'main': None, 'stats': self.stats}

        except (FileOperationError, ComponentExecutionError):
            # Re-raise custom exceptions
            raise
        except Exception as e:
            logger.error(f"[{self.id}] Processing failed: {e}")
            raise ComponentExecutionError(self.id, f"Excel output processing failed: {e}", e) from e

    # ------------------------------------------------------------------
    # Date Pattern and Decimal Precision Formatting
    # ------------------------------------------------------------------

    def _apply_date_patterns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Format datetime columns according to per-column ``date_pattern``.

        Walks ``self.input_schema`` and, for each column declared as a
        datetime with a ``date_pattern``, converts the column to a string
        formatted with that pattern. NaT values become empty strings.

        Args:
            df: Working copy DataFrame to format in place.

        Returns:
            The same DataFrame with datetime columns replaced by formatted strings.
        """
        schema = getattr(self, "input_schema", None) or []
        if not schema:
            return df

        for col in schema:
            if not isinstance(col, dict):
                continue
            name = col.get("name")
            if not name or name not in df.columns:
                continue
            col_type = (col.get("type") or "").lower()
            pattern = col.get("date_pattern") or ""
            if not pattern or col_type not in ("date", "datetime"):
                continue
            series = df[name]
            if not pd.api.types.is_datetime64_any_dtype(series):
                try:
                    # Pass the schema's date_pattern as ``format=`` so pandas
                    # uses the vectorised C parser instead of dateutil
                    # (which emits a "Could not infer format" UserWarning).
                    series = pd.to_datetime(series, format=pattern, errors="coerce")
                except Exception:
                    logger.warning(
                        f"[{self.id}] Column '{name}' could not be coerced "
                        f"to datetime; skipping date_pattern formatting."
                    )
                    continue
            formatted = series.dt.strftime(pattern)
            df[name] = formatted.where(series.notna(), "")
        return df

    def _build_col_formats(self) -> dict:
        """Build a mapping of column name -> openpyxl number format string.

        For Decimal/float columns with a ``precision`` defined in the schema,
        returns an Excel number format like ``"0.0000000000"`` (precision zeros
        after the decimal point) so the cell displays exactly that many decimal
        places.  

        Returns:
            Dict mapping column name to number format string.
        """
        schema = getattr(self, "input_schema", None) or []
        col_formats = {}
        for col in schema:
            if not isinstance(col, dict):
                continue
            name = col.get("name")
            if not name:
                continue
            col_type = (col.get("type") or "").lower()
            if col_type not in ("decimal", "bigdecimal", "numeric", "number", "float"):
                continue
            precision = col.get("precision")
            if precision is None:
                continue
            p = int(precision)
            if p < 0:
                continue
            col_formats[name] = ("0." + "0" * p) if p > 0 else "0"
        return col_formats